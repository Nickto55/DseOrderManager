from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

class ExcelDataInserter:
    """
    Класс для вставки вложенной структуры данных в существующий Excel-файл.
    Создаёт новый лист с заданным названием и записывает данные в табличном виде.
    """

    def __init__(self, file_path):
        """
        :param file_path: Путь к существующему Excel-файлу
        """
        self.file_path = file_path
        self.wb = load_workbook(file_path)

        self.fill_color1 = PatternFill(start_color="6f747c", fill_type="solid")

    def insert_data(self, data, sheet_name, headers=None):
        """
        Вставляет данные в новый лист Excel-файла.
        
        :param data: Вложенный словарь с данными (ваша структура)
        :param sheet_name: Название нового листа 
                           (если существует — будет удалён и создан заново)
        :param headers: Список заголовков столбцов (опционально, 
                        извлекаются автоматически из данных)
        :return: Путь к сохранённому файлу
        """
        # Удаляем лист, если он уже существует
        if sheet_name in self.wb.sheetnames:
            del self.wb[sheet_name]

        # Создаём новый лист
        ws = self.wb.create_sheet(title=sheet_name)

        # Определяем заголовки из структуры данных, если не переданы явно
        if headers is None:
            first_item = None
            for product_key, product_value in data.items():
                for detail_key, detail_value in product_value.items():
                    first_item = detail_value
                    break
                if first_item:
                    break
            headers = list(first_item.keys()) if first_item else []

        # Записываем заголовки (первая строка)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.fill_color1
            cell.font = Font(color="f2ecde")

        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"


        # Записываем данные
        row_idx = 2
        for product_key, product_value in data.items():
            for detail_key, detail_value in product_value.items():
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=detail_value.get(header, ''))
                row_idx += 1

        # Автоширина столбцов
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        self.wb.freeze_panes = 'A2'
        # Сохраняем файл
        self.wb.save(self.file_path)
        return self.file_path

    def close(self):
        """Закрывает workbook."""
        self.wb.close()
